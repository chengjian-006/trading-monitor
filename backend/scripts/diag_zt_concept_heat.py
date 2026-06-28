"""今日涨停板·按炒作概念热度聚合 + Excel 导出 (只读诊断)。

口径:
  封板池 limit_up_pool(info)   -> 上板家数/最高连板, 单只票按优先级归一到一个主概念。
  炸板池 open_limit_pool(info) -> 无官方题材, 按个股主业人工归类, 用于估算概念封板率。
  封板率(估) = 该概念封板家数 / (封板家数 + 该概念炸板家数)。

跑法: py -3 -m backend.scripts.diag_zt_concept_heat [YYYYMMDD]
"""
import asyncio
import re
import sys

import httpx

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

H = {"User-Agent": "Mozilla/5.0", "Referer": "https://data.10jqka.com.cn/"}

# 主概念归一规则: (概念名, 关键词列表), 自上而下首个命中即归入, 故越靠前=优先级越高。
THEME_RULES = [
    ("PCB·覆铜板·铜箔", ["覆铜板", "PCB", "铜箔", "电子布", "PET铜箔", "HDI", "PI膜", "PI薄膜", "钻针", "电极箔"]),
    ("玻璃基板·先进封装·面板", ["玻璃基板", "玻璃基封", "基板玻璃", "TGV", "UTG", "先进封装",
                                 "封装基板", "FCBGA", "载板", "Micro-LED", "面板", "MED", "光伏玻璃", "浮法玻璃"]),
    ("半导体设备·存储芯片", ["存储", "半导体设备", "HBM", "MLCC", "电子陶瓷", "晶圆", "SPM",
                             "洁净室", "电子化学品", "半导体", "芯片"]),
    ("AI算力·液冷·光模块·IDC", ["算力", "液冷", "数据中心", "光模块", "智算", "服务器", "IDC", "光互连"]),
    ("人形机器人·机器人", ["机器人", "人形"]),
    ("商业航天·3D打印·军工", ["商业航天", "3D打印", "航天", "军工", "无人机", "航空装备", "VR/XR"]),
    ("核电·可控核聚变", ["可控核聚变", "核聚变", "核电", "中核"]),
    ("固态电池·锂电材料", ["固态电池", "锂电", "电解液", "隔膜", "芳纶"]),
    ("光纤·特高压", ["光纤", "光缆", "特高压"]),
    ("央企·中字头·国资重组", ["中字头", "央企", "国资", "国企", "资产重组", "资产置换", "控制权", "重整"]),
]

# 炸板个股按主业归类 (code -> 概念名); 未列出者归"其他/独立"。
BROKEN_THEME = {
    "603175": "PCB·覆铜板·铜箔", "002436": "PCB·覆铜板·铜箔", "301150": "PCB·覆铜板·铜箔",
    "002160": "PCB·覆铜板·铜箔", "002484": "AI算力·液冷·光模块·IDC",
    "300196": "玻璃基板·先进封装·面板", "605006": "玻璃基板·先进封装·面板",
    "002066": "玻璃基板·先进封装·面板", "002976": "人形机器人·机器人",
}


def board_n(it):
    hv = it.get("high_days_value")
    if hv:
        try:
            return int(hv) >> 16
        except Exception:
            pass
    nums = re.findall(r"\d+", str(it.get("high_days") or ""))
    return int(nums[-1]) if nums else 1


def classify(reason: str) -> str:
    for theme, kws in THEME_RULES:
        if any(k in reason for k in kws):
            return theme
    return "其他·独立个股"


async def fetch(client, ep, order=""):
    url = (f"https://data.10jqka.com.cn/dataapi/limit_up/{ep}"
           f"?page=1&limit=200&field=199112,10,9001,330323,330324,330325,"
           f"9002,330329,133971,133970,1968584,3475914,9003,9004{order}&date=20260617")
    r = await client.get(url, headers=H)
    return (r.json() or {}).get("data", {}).get("info") or []


async def main():
    date = sys.argv[1] if len(sys.argv) > 1 else "20260617"
    async with httpx.AsyncClient(timeout=15, follow_redirects=True, trust_env=False) as c:
        sealed = await fetch(c, "limit_up_pool", "&order_field=330323&order_type=0")
        broken = await fetch(c, "open_limit_pool")

    agg = {}  # theme -> {"stocks":[(name,height,reason)], "broken":int}
    for it in sealed:
        th = classify(it.get("reason_type") or "")
        d = agg.setdefault(th, {"stocks": [], "broken": 0})
        d["stocks"].append((it.get("name"), board_n(it), it.get("reason_type") or ""))
    for it in broken:
        th = BROKEN_THEME.get(str(it.get("code")), "其他·独立个股")
        agg.setdefault(th, {"stocks": [], "broken": 0})["broken"] += 1

    total = len(sealed)
    rows = []
    for th, d in agg.items():
        n = len(d["stocks"])
        if n == 0:
            continue
        hi = max((s[1] for s in d["stocks"]), default=1)
        leaders = sorted(d["stocks"], key=lambda s: -s[1])[:3]
        sr = n / (n + d["broken"]) if (n + d["broken"]) else 1.0
        rows.append({"theme": th, "n": n, "pct": n / total, "hi": hi, "broken": d["broken"],
                     "seal_rate": sr,
                     "leaders": "、".join(f"{s[0]}({s[1]}板)" if s[1] >= 2 else s[0] for s in leaders)})
    rows.sort(key=lambda r: (-r["n"], -r["hi"]))

    print(f"涨停板概念热度  date={date}  封板{total} 炸板{len(broken)}")
    print(f"{'概念':<26}{'家数':<5}{'占比':<7}{'最高板':<7}{'封板率估':<9}龙头")
    print("-" * 100)
    for r in rows:
        print(f"{r['theme']:<24}{r['n']:<5}{r['pct']*100:>5.1f}%  {r['hi']}板    "
              f"{r['seal_rate']*100:>5.1f}%   {r['leaders']}")

    # ---- Excel 导出 ----
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "概念热度"
        hdr = ["排名", "炒作概念", "上板家数", "占全场", "最高连板", "炸板数", "封板率(估)", "代表龙头"]
        ws.append(hdr)
        for i, r in enumerate(rows, 1):
            ws.append([i, r["theme"], r["n"], round(r["pct"], 4), f"{r['hi']}板",
                       r["broken"], round(r["seal_rate"], 4), r["leaders"]])
        # 样式
        head_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ("D", "G"):
            for cell in ws[col][1:]:
                cell.number_format = "0.0%"
        widths = {"A": 6, "B": 26, "C": 10, "D": 9, "E": 10, "F": 8, "G": 11, "H": 60}
        for c, w in widths.items():
            ws.column_dimensions[c].width = w
        # 明细页
        ws2 = wb.create_sheet("封板个股明细")
        ws2.append(["�X概念", "名称", "连板高度", "涨停原因"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
        for r in rows:
            d = agg[r["theme"]]
            for nm, h, rs in sorted(d["stocks"], key=lambda s: -s[1]):
                ws2.append([r["theme"], nm, f"{h}板", rs])
        for c, w in {"A": 26, "B": 12, "C": 10, "D": 50}.items():
            ws2.column_dimensions[c].width = w
        out = f"D:/财务管理/交易系统/trading-monitor/涨停板概念热度_{date}.xlsx"
        wb.save(out)
        print(f"\n已导出 Excel: {out}")
    except ImportError:
        print("\n(openpyxl 未安装, 跳过 Excel 导出)")


if __name__ == "__main__":
    asyncio.run(main())
